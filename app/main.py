import os
from fastapi import FastAPI, HTTPException, Query, Depends, Header
from fastapi.middleware.cors import CORSMiddleware
import yfinance as yf
import openpyxl
from typing import Optional



# Firebase Admin SDK
import firebase_admin
from firebase_admin import credentials, auth as firebase_auth, firestore

# Initialize Firebase Admin
cred = credentials.Certificate("serviceAccountKey.json")
firebase_admin.initialize_app(cred)

db = firestore.client()

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://stocks-d4bba.web.app/"],  # In production, restrict this to your frontend domain
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def get_current_user(authorization: Optional[str] = Header(None)):
    if not authorization:
        raise HTTPException(status_code=401, detail="No authorization header")
    parts = authorization.split(" ")
    if len(parts) != 2 or parts[0].lower() != "bearer":
        raise HTTPException(status_code=401, detail="Invalid token type")
    id_token = parts[1]
    try:
        decoded_token = firebase_auth.verify_id_token(id_token)
        email = decoded_token.get("email")
        if not email:
            raise HTTPException(status_code=401, detail="Invalid token: no email")
        # Ensure user document exists in Firestore
        user_ref = db.collection("users").document(email)
        if not user_ref.get().exists:
            user_ref.set({"stocks": []})
        return email
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid or expired token")

def get_simulated_american_bull_data(stock_code):
    stock_data = specific_stock(stock_code)
    if not stock_data:
        return {"Date": "N/A", "Signal": "N/A"}
    latest_entry = stock_data[0]
    return {"Date": latest_entry.get("Date", "N/A"), "Signal": latest_entry.get("Signal", "N/A")}

@app.post("/add_stock")
def add_stock(stock_symbol: str = Query(...), user_email: str = Depends(get_current_user)):
    stock_symbol = stock_symbol.upper()
    ticker = yf.Ticker(stock_symbol)
    try:
        data = ticker.history(period='1d')
        if data.empty:
            raise ValueError("Invalid stock symbol")
    except:
        raise HTTPException(status_code=400, detail=f"Invalid stock symbol: {stock_symbol}")

    user_ref = db.collection("users").document(user_email)
    user_data = user_ref.get().to_dict()
    stocks = user_data.get("stocks", [])
    if stock_symbol in stocks:
        return {"message": f"Stock {stock_symbol} is already in your list."}
    # Add the new stock
    stocks.append(stock_symbol)
    user_ref.update({"stocks": stocks})
    return {"message": f"Stock {stock_symbol} added to your list."}

@app.post("/remove_stock")
def remove_stock(stock_symbol: str = Query(...), user_email: str = Depends(get_current_user)):
    stock_symbol = stock_symbol.upper()
    user_ref = db.collection("users").document(user_email)
    user_data = user_ref.get().to_dict()
    stocks = user_data.get("stocks", [])

    if stock_symbol not in stocks:
        return {"message": f"Stock {stock_symbol} is not in your list."}

    # Remove the stock
    stocks = [s for s in stocks if s != stock_symbol]
    user_ref.update({"stocks": stocks})

    # Remove stock data from Excel sheets
    file_name = "trade_log.xlsx"
    if os.path.exists(file_name):
        wb = openpyxl.load_workbook(file_name)
        for sheet_name in ["Stock Analysis", "American Bull Info"]:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_to_delete = [
                    row[0].row for row in ws.iter_rows(min_row=2)
                    if row[0].value == stock_symbol
                ]
                for row_num in sorted(rows_to_delete, reverse=True):
                    ws.delete_rows(row_num)
        wb.save(file_name)

    return {"message": f"Stock {stock_symbol} removed and data deleted."}

@app.post("/remove_stock_data")
def remove_stock_data(stock_symbol: str = Query(...), user_email: str = Depends(get_current_user)):
    stock_symbol = stock_symbol.upper()
    file_name = "trade_log.xlsx"
    if not os.path.exists(file_name):
        return {"error": "Data file not found."}

    wb = openpyxl.load_workbook(file_name)
    for sheet_name in ["Stock Analysis", "American Bull Info"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_to_delete = [
                row[0].row for row in ws.iter_rows(min_row=2)
                if row[0].value == stock_symbol
            ]
            for row_num in sorted(rows_to_delete, reverse=True):
                ws.delete_rows(row_num)
    wb.save(file_name)
    return {"message": f"Data for stock {stock_symbol} has been removed."}

@app.get("/stocks")
def get_stocks(user_email: str = Depends(get_current_user)):
    user_ref = db.collection("users").document(user_email)
    user_data = user_ref.get().to_dict()
    return {"stocks": user_data.get("stocks", [])}

@app.post("/execute_analysis")
def execute_analysis(user_email: str = Depends(get_current_user)):
    user_ref = db.collection("users").document(user_email)
    user_data = user_ref.get().to_dict()
    user_stocks = user_data.get("stocks", [])
    if not user_stocks:
        return {"status": "No stocks to analyze", "results": []}

    initialize_excel()
    results = []
    today_date = get_todays_date()

    for stock in user_stocks:
        stock = stock.upper()
        bull_data = get_simulated_american_bull_data(stock)
        bull_date, bull_signal = bull_data["Date"], bull_data["Signal"]

        log_american_bull_info(stock, bull_signal, bull_date)

        latest_k, latest_d = calculate_stochastic(stock)
        if latest_k is None or latest_d is None:
            message = f"Could not calculate stochastic oscillator for {stock}"
            results.append({"stock": stock, "message": message})
            continue

        zone = determine_zone(latest_k, latest_d)
        decision = decide_action(zone)

        try:
            ticker = yf.Ticker(stock)
            latest_price = ticker.history(period='1d')['Close'].iloc[-1]
        except Exception as e:
            message = f"Error getting price for {stock}: {e}"
            results.append({"stock": stock, "message": message})
            continue

        log_stock_analysis(stock, latest_price, latest_k, latest_d, zone, decision)

        if bull_date == today_date:
            if "BUY" in bull_signal and "Green Zone" in zone:
                action = "Buy"
                shares = round(250 / latest_price, 6)
                executed = True
                log_trade(action, stock, latest_price, shares, executed, zone, latest_k)
                message = f"Simulated buying {shares} shares of {stock} at {latest_price}"
                results.append({"stock": stock, "action": action, "message": message})
            elif "SELL" in bull_signal and "Red Zone" in zone:
                action = "Sell"
                shares_owned = 10  # Simulate some shares
                executed = True
                log_trade(action, stock, latest_price, shares_owned, executed, zone, latest_k)
                message = f"Simulated selling {shares_owned} shares of {stock} at {latest_price}"
                results.append({"stock": stock, "action": action, "message": message})
            else:
                message = f"No action for {stock}: Signal and zone do not match."
                results.append({"stock": stock, "message": message})
        else:
            message = f"No action for {stock}: Signal date does not match today."
            results.append({"stock": stock, "message": message})

    return {"status": "Analysis executed", "results": results}

@app.get("/data/{sheet_name}")
def get_data(sheet_name: str, user_email: str = Depends(get_current_user)):
    file_name = "trade_log.xlsx"
    if not os.path.exists(file_name):
        return {"error": "Data file not found."}

    wb = openpyxl.load_workbook(file_name, data_only=True)
    if sheet_name not in wb.sheetnames:
        return {"error": f"Sheet {sheet_name} not found."}

    ws = wb[sheet_name]
    data = []
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        data.append(record)
    return {"data": data}



### Scraping
import requests
from bs4 import BeautifulSoup

def recommend_data():
    """Scrape recommended stock data from American Bulls website."""
    url = "https://www.americanbulls.com/Default.aspx?lang=en"

    response = requests.get(url)
    soup = BeautifulSoup(response.content, "lxml")

    # Find the table containing stock data
    table = soup.find("table", class_="table-hover")
    if not table:
        print("No table found on the page.")
        return []

    rows = table.find_all("tr", class_="gridRows")
    stock_data = []

    # Helper function to clean text
    def clean_text(text):
        return ' '.join(text.split()).strip()

    for row in rows:
        date = row.find("div", id=lambda x: x and "Tarih" in x)
        stock_name = row.find("a", class_="dxbs-hyperlink")
        signal = row.find("div", id=lambda x: x and "gridlevel" in x)
        buy_level = row.find("div", id=lambda x: x and "gridprice" in x)
        close_price = row.find("div", id=lambda x: x and "gridclose" in x)

        stock_data.append({
            "Date": clean_text(date.text) if date else "N/A",
            "Stock Name": clean_text(stock_name.text) if stock_name else "N/A",
            "Signal": clean_text(signal.text) if signal else "N/A",
            "Buy Level": clean_text(buy_level.text) if buy_level else "N/A",
            "Close Price": clean_text(close_price.text) if close_price else "N/A",
        })
    return stock_data

def specific_stock(stock_code):
    """Scrape specific stock data from American Bulls website."""
    url = f"https://www.americanbulls.com/SignalPage.aspx?lang=en&Ticker={stock_code}"
    response = requests.get(url)
    soup = BeautifulSoup(response.content, "lxml")

    table = soup.find("table", id="Content_SignalHistory_SignalShortHistoryGrid_DXMainTable")
    if table is None:
        print(f"Table not found for stock: {stock_code}")
        return []

    table_data = []
    for row in table.find("tbody").find_all("tr"):
        cells = row.find_all("td")
        if cells:
            table_data.append({
                "Date": cells[0].text.strip() if len(cells) > 0 else "N/A",
                "Price": cells[1].text.strip() if len(cells) > 1 else "N/A",
                "Signal": cells[2].text.strip() if len(cells) > 2 else "N/A",
                "Change%": cells[3].text.strip() if len(cells) > 3 else "N/A",
                "Value": cells[4].text.strip() if len(cells) > 4 else "N/A",
            })
    return table_data

# Example usage:
if __name__ == "__main__":
    data = specific_stock("IRS")
    print(data)


import os
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import yfinance as yf

def initialize_excel(file_name="trade_log.xlsx"):
    """Initialize Excel file with necessary sheets and headers."""
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Trade Log"
        ws.append(["Date", "Action", "Stock Name", "Price", "Shares", "Zone", "Percent", "Executed"])

        ws_analysis = wb.create_sheet(title="Stock Analysis")
        ws_analysis.append(["Stock Name", "Price", "%K", "%D", "Zone", "Decision"])

        ws_bull = wb.create_sheet(title="American Bull Info")
        ws_bull.append(["Stock Name", "Signal", "Date"])

        wb.save(file_name)
    else:
        wb = openpyxl.load_workbook(file_name)
        if "Stock Analysis" not in wb.sheetnames:
            ws_analysis = wb.create_sheet(title="Stock Analysis")
            ws_analysis.append(["Stock Name", "Price", "%K", "%D", "Zone", "Decision"])
        if "American Bull Info" not in wb.sheetnames:
            ws_bull = wb.create_sheet(title="American Bull Info")
            ws_bull.append(["Stock Name", "Signal", "Date"])
        wb.save(file_name)
    return wb

def calculate_stochastic(stock_symbol, period=14):
    """Calculate the stochastic oscillator for a given stock."""
    try:
        data = yf.download(stock_symbol, period='1mo', interval='1d')
        if data.empty:
            print(f"No data for {stock_symbol}")
            return None, None

        df = data.copy()
        df['Lowest Low'] = df['Low'].rolling(window=period).min()
        df['Highest High'] = df['High'].rolling(window=period).max()
        df['%K'] = ((df['Close'] - df['Lowest Low']) / (df['Highest High'] - df['Lowest Low'])) * 100
        df['%D'] = df['%K'].rolling(window=3).mean()

        latest_k = df['%K'].iloc[-1]
        latest_d = df['%D'].iloc[-1]
        return latest_k, latest_d
    except Exception as e:
        print(f"Error calculating stochastic for {stock_symbol}: {e}")
        return None, None

def determine_zone(latest_k, latest_d):
    """Determine the zone based on %K and %D values."""
    if latest_k is None or latest_d is None:
        return "Unknown"
    if latest_k > 80 or latest_d > 80:
        return "Red Zone: Overbought - Potential Sell Opportunity"
    elif latest_k < 20 or latest_d < 20:
        return "Green Zone: Oversold - Potential Buy Opportunity"
    else:
        return "Neutral Zone: Hold Phase"

def decide_action(zone):
    """Decide action based on the zone."""
    if "Overbought" in zone:
        return "Consider Selling"
    elif "Oversold" in zone:
        return "Consider Buying"
    else:
        return "Hold"

def log_stock_analysis(stock_name, price, latest_k, latest_d, zone, decision, file_name="trade_log.xlsx"):
    """Log stock analysis data into Excel sheet."""
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Stock Analysis"]
    stock_found = False
    for row in ws.iter_rows(min_row=2):
        if row[0].value == stock_name:
            row[1].value = price
            row[2].value = latest_k
            row[3].value = latest_d
            row[4].value = zone
            row[5].value = decision
            stock_found = True
            break
    if not stock_found:
        ws.append([stock_name, price, latest_k, latest_d, zone, decision])
    wb.save(file_name)

def log_american_bull_info(stock_name, signal, date, file_name="trade_log.xlsx"):
    """Log American Bull signal information into Excel sheet."""
    wb = openpyxl.load_workbook(file_name)
    ws = wb["American Bull Info"]
    stock_found = False
    for row in ws.iter_rows(min_row=2):
        if row[0].value == stock_name:
            row[1].value = signal
            row[2].value = date
            stock_found = True
            break
    if not stock_found:
        ws.append([stock_name, signal, date])
    wb.save(file_name)

def log_trade(action, stock_name, price, shares, executed, zone, percent_value, file_name="trade_log.xlsx"):
    """Log trade action into Excel sheet."""
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Trade Log"]
    today_date = get_todays_date()
    executed_status = "Yes" if executed else "No"
    ws.append([today_date, action, stock_name, price, shares, zone, percent_value, executed_status])
    wb.save(file_name)

def get_todays_date():
    """Get today's date in mm/dd format."""
    return datetime.today().strftime('%m/%d')
