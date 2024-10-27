import os
from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
import uvicorn
import yfinance as yf
import numpy as np
import openpyxl

# Import functions from stock_data.py
from stock_data import (
    initialize_excel,
    calculate_stochastic,
    determine_zone,
    decide_action,
    log_stock_analysis,
    log_trade,
    get_todays_date,
    log_american_bull_info,
)

app = FastAPI()

# Configure CORS middleware (adjust allowed origins in production)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Replace with your frontend URL in production
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# In-memory storage for stocks and credentials
user_stocks = []
rh_credentials = {"username": "", "password": ""}

@app.post("/add_stock")
def add_stock(stock_symbol: str = Query(...)):
    """Add a stock to the user's list after validation."""
    stock_symbol = stock_symbol.upper()
    ticker = yf.Ticker(stock_symbol)
    try:
        data = ticker.history(period='1d')
        if data.empty:
            raise ValueError("Invalid stock symbol")
    except Exception:
        raise HTTPException(status_code=400, detail=f"Invalid stock symbol: {stock_symbol}")

    if stock_symbol not in user_stocks:
        user_stocks.append(stock_symbol)
        return {"message": f"Stock {stock_symbol} added to your list."}
    else:
        return {"message": f"Stock {stock_symbol} is already in your list."}

@app.post("/remove_stock")
def remove_stock(stock_symbol: str = Query(...)):
    """Remove a stock from the user's list and delete its data."""
    stock_symbol = stock_symbol.upper()
    if stock_symbol in user_stocks:
        user_stocks.remove(stock_symbol)
        # Remove stock data from Excel sheets
        file_name = "trade_log.xlsx"
        if os.path.exists(file_name):
            wb = openpyxl.load_workbook(file_name)
            for sheet_name in ["Stock Analysis", "American Bull Info"]:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows_to_delete = [
                        row[0].row for row in ws.iter_rows(min_row=2)
                        if row[0].value == stock_symbol
                    ]
                    for row_num in sorted(rows_to_delete, reverse=True):
                        ws.delete_rows(row_num)
            wb.save(file_name)
        return {"message": f"Stock {stock_symbol} removed and data deleted."}
    else:
        return {"message": f"Stock {stock_symbol} is not in your list."}

@app.post("/remove_stock_data")
def remove_stock_data(stock_symbol: str = Query(...)):
    """Remove stock data from Excel sheets without removing from the stock list."""
    stock_symbol = stock_symbol.upper()
    file_name = "trade_log.xlsx"
    if not os.path.exists(file_name):
        return {"error": "Data file not found."}

    wb = openpyxl.load_workbook(file_name)
    for sheet_name in ["Stock Analysis", "American Bull Info"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_to_delete = [
                row[0].row for row in ws.iter_rows(min_row=2)
                if row[0].value == stock_symbol
            ]
            for row_num in sorted(rows_to_delete, reverse=True):
                ws.delete_rows(row_num)
    wb.save(file_name)
    return {"message": f"Data for stock {stock_symbol} has been removed."}

@app.get("/stocks")
def get_stocks():
    """Get the list of user's stocks."""
    return {"stocks": user_stocks}

def get_simulated_american_bull_data(stock_code):
    """Simulate American Bull data for a stock."""
    simulated_signal = "BUY" if np.random.rand() > 0.5 else "SELL"
    bull_date = get_todays_date()
    return {"Date": bull_date, "Signal": simulated_signal}

@app.post("/execute_analysis")
def execute_analysis():
    """Execute analysis on the user's stock list."""
    initialize_excel()
    results = []
    today_date = get_todays_date()

    for stock in user_stocks:
        stock = stock.upper()
        # Simulate American Bull info
        bull_data = get_simulated_american_bull_data(stock)
        bull_date, bull_signal = bull_data["Date"], bull_data["Signal"]

        # Log American Bull info
        log_american_bull_info(stock, bull_signal, bull_date)

        # Calculate stochastic oscillator
        latest_k, latest_d = calculate_stochastic(stock)
        if latest_k is None or latest_d is None:
            message = f"Could not calculate stochastic oscillator for {stock}"
            results.append({"stock": stock, "message": message})
            continue

        # Determine the zone and decide action
        zone = determine_zone(latest_k, latest_d)
        decision = decide_action(zone)

        # Get latest price
        try:
            ticker = yf.Ticker(stock)
            latest_price = ticker.history(period='1d')['Close'].iloc[-1]
        except Exception as e:
            message = f"Error getting price for {stock}: {e}"
            results.append({"stock": stock, "message": message})
            continue

        # Log stock analysis
        log_stock_analysis(stock, latest_price, latest_k, latest_d, zone, decision)

        # Simulate trade action based on American Bull info and zone
        if bull_date == today_date:
            if "BUY" in bull_signal and "Green Zone" in zone:
                action = "Buy"
                shares = round(250 / latest_price, 6)
                executed = True
                log_trade(action, stock, latest_price, shares, executed, zone, latest_k)
                message = f"Simulated buying {shares} shares of {stock} at {latest_price}"
                results.append({"stock": stock, "action": action, "message": message})
            elif "SELL" in bull_signal and "Red Zone" in zone:
                action = "Sell"
                shares_owned = 10  # Simulated shares owned
                executed = True
                log_trade(action, stock, latest_price, shares_owned, executed, zone, latest_k)
                message = f"Simulated selling {shares_owned} shares of {stock} at {latest_price}"
                results.append({"stock": stock, "action": action, "message": message})
            else:
                message = f"No action for {stock}: Signal and zone do not match."
                results.append({"stock": stock, "message": message})
        else:
            message = f"No action for {stock}: Signal date does not match today."
            results.append({"stock": stock, "message": message})

    return {"status": "Analysis executed", "results": results}

@app.get("/data/{sheet_name}")
def get_data(sheet_name: str):
    """Fetch data from the specified Excel sheet."""
    file_name = "trade_log.xlsx"
    if not os.path.exists(file_name):
        return {"error": "Data file not found."}

    wb = openpyxl.load_workbook(file_name, data_only=True)
    if sheet_name not in wb.sheetnames:
        return {"error": f"Sheet {sheet_name} not found."}

    ws = wb[sheet_name]
    data = []
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        data.append(record)
    return {"data": data}

@app.post("/save_robinhood_credentials")
def save_robinhood_credentials(
    rh_username: str = Query(...), rh_password: str = Query(...)
):
    """Save Robinhood credentials (in-memory for this example)."""
    rh_credentials["username"] = rh_username
    rh_credentials["password"] = rh_password
    return {"message": "Robinhood credentials saved successfully."}

if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
