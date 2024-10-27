import os
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import yfinance as yf

def initialize_excel(file_name="trade_log.xlsx"):
    """Initialize Excel file with necessary sheets and headers."""
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Trade Log"
        ws.append(["Date", "Action", "Stock Name", "Price", "Shares", "Zone", "Percent", "Executed"])

        ws_analysis = wb.create_sheet(title="Stock Analysis")
        ws_analysis.append(["Stock Name", "Price", "%K", "%D", "Zone", "Decision"])

        ws_bull = wb.create_sheet(title="American Bull Info")
        ws_bull.append(["Stock Name", "Signal", "Date"])

        wb.save(file_name)
    else:
        wb = openpyxl.load_workbook(file_name)
        if "Stock Analysis" not in wb.sheetnames:
            ws_analysis = wb.create_sheet(title="Stock Analysis")
            ws_analysis.append(["Stock Name", "Price", "%K", "%D", "Zone", "Decision"])
        if "American Bull Info" not in wb.sheetnames:
            ws_bull = wb.create_sheet(title="American Bull Info")
            ws_bull.append(["Stock Name", "Signal", "Date"])
        wb.save(file_name)
    return wb

def calculate_stochastic(stock_symbol, period=14):
    """Calculate the stochastic oscillator for a given stock."""
    try:
        data = yf.download(stock_symbol, period='1mo', interval='1d')
        if data.empty:
            print(f"No data for {stock_symbol}")
            return None, None

        df = data.copy()
        df['Lowest Low'] = df['Low'].rolling(window=period).min()
        df['Highest High'] = df['High'].rolling(window=period).max()
        df['%K'] = ((df['Close'] - df['Lowest Low']) / (df['Highest High'] - df['Lowest Low'])) * 100
        df['%D'] = df['%K'].rolling(window=3).mean()

        latest_k = df['%K'].iloc[-1]
        latest_d = df['%D'].iloc[-1]
        return latest_k, latest_d
    except Exception as e:
        print(f"Error calculating stochastic for {stock_symbol}: {e}")
        return None, None

def determine_zone(latest_k, latest_d):
    """Determine the zone based on %K and %D values."""
    if latest_k is None or latest_d is None:
        return "Unknown"
    if latest_k > 80 or latest_d > 80:
        return "Red Zone: Overbought - Potential Sell Opportunity"
    elif latest_k < 20 or latest_d < 20:
        return "Green Zone: Oversold - Potential Buy Opportunity"
    else:
        return "Neutral Zone: Hold Phase"

def decide_action(zone):
    """Decide action based on the zone."""
    if "Overbought" in zone:
        return "Consider Selling"
    elif "Oversold" in zone:
        return "Consider Buying"
    else:
        return "Hold"

def log_stock_analysis(stock_name, price, latest_k, latest_d, zone, decision, file_name="trade_log.xlsx"):
    """Log stock analysis data into Excel sheet."""
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Stock Analysis"]
    stock_found = False
    for row in ws.iter_rows(min_row=2):
        if row[0].value == stock_name:
            row[1].value = price
            row[2].value = latest_k
            row[3].value = latest_d
            row[4].value = zone
            row[5].value = decision
            stock_found = True
            break
    if not stock_found:
        ws.append([stock_name, price, latest_k, latest_d, zone, decision])
    wb.save(file_name)

def log_american_bull_info(stock_name, signal, date, file_name="trade_log.xlsx"):
    """Log American Bull signal information into Excel sheet."""
    wb = openpyxl.load_workbook(file_name)
    ws = wb["American Bull Info"]
    stock_found = False
    for row in ws.iter_rows(min_row=2):
        if row[0].value == stock_name:
            row[1].value = signal
            row[2].value = date
            stock_found = True
            break
    if not stock_found:
        ws.append([stock_name, signal, date])
    wb.save(file_name)

def log_trade(action, stock_name, price, shares, executed, zone, percent_value, file_name="trade_log.xlsx"):
    """Log trade action into Excel sheet."""
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Trade Log"]
    today_date = get_todays_date()
    executed_status = "Yes" if executed else "No"
    ws.append([today_date, action, stock_name, price, shares, zone, percent_value, executed_status])
    wb.save(file_name)

def get_todays_date():
    """Get today's date in mm/dd format."""
    return datetime.today().strftime('%m/%d')
