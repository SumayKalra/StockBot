import openpyxl
from openpyxl import Workbook
import os
from robin_stocks.robinhood import authentication as r
from robin_stocks.robinhood import account
from robin_stocks.robinhood import orders
from robin_stocks.robinhood import stocks
from ScrapingDetails import recommmend_data, specifc_stock
from datetime import datetime
import pandas as pd
import numpy as np

# Initialize Excel File
def initialize_excel(file_name="trade_log.xlsx"):
    if not os.path.exists(file_name):
        # Create a new workbook and add headers if the file doesn't exist
        wb = Workbook()
        ws = wb.active
        ws.title = "Trade Log"
        ws.append(["Date", "Action", "Stock Name", "Price", "Shares", "Zone", "Percent", "Executed"])
        
        # Create the Stock Analysis sheet
        ws_analysis = wb.create_sheet(title="Stock Analysis")
        ws_analysis.append(["Stock Name", "Price", "%K", "%D", "Zone", "Decision"])

        # Create the American Bull Info sheet
        ws_bull = wb.create_sheet(title="American Bull Info")
        ws_bull.append(["Stock Name", "Signal", "Date"])

        wb.save(file_name)
    else:
        # Open the existing workbook
        wb = openpyxl.load_workbook(file_name)
        ws = wb["Trade Log"]
        # Check and add missing columns
        headers = [cell.value for cell in ws[1]]
        if "Zone" not in headers:
            ws.cell(row=1, column=len(headers)+1, value="Zone")
            headers.append("Zone")
        if "Percent" not in headers:
            ws.cell(row=1, column=len(headers)+1, value="Percent")
            headers.append("Percent")
        if "Executed" not in headers:
            ws.cell(row=1, column=len(headers)+1, value="Executed")
            headers.append("Executed")
        # Check if the "Stock Analysis" sheet exists
        if "Stock Analysis" not in wb.sheetnames:
            ws_analysis = wb.create_sheet(title="Stock Analysis")
            ws_analysis.append(["Stock Name", "Price", "%K", "%D", "Zone", "Decision"])
        # Check if the "American Bull Info" sheet exists
        if "American Bull Info" not in wb.sheetnames:
            ws_bull = wb.create_sheet(title="American Bull Info")
            ws_bull.append(["Stock Name", "Signal", "Date"])
        wb.save(file_name)
    return wb

# Function to calculate the stochastic oscillator
def calculate_stochastic(stock_symbol, period=14):
    # Get historical data
    historicals = stocks.get_stock_historicals(stock_symbol, interval='day', span='month')
    if not historicals:
        print(f"Could not retrieve historical data for {stock_symbol}.")
        return None, None

    # Convert to DataFrame
    df = pd.DataFrame(historicals)
    df['close_price'] = df['close_price'].astype(float)
    df['high_price'] = df['high_price'].astype(float)
    df['low_price'] = df['low_price'].astype(float)

    # Calculate %K
    df['Lowest Low'] = df['low_price'].rolling(window=period).min()
    df['Highest High'] = df['high_price'].rolling(window=period).max()
    df['%K'] = ((df['close_price'] - df['Lowest Low']) / (df['Highest High'] - df['Lowest Low'])) * 100
    # Calculate %D
    df['%D'] = df['%K'].rolling(window=3).mean()

    # Get the latest %K and %D values
    latest_k = df['%K'].iloc[-1]
    latest_d = df['%D'].iloc[-1]

    return latest_k, latest_d

# Function to determine the zone
def determine_zone(latest_k, latest_d):
    if latest_k is None or latest_d is None:
        return "Unknown"
    elif latest_k > 80 or latest_d > 80:
        return "Red Zone: Overbought - Potential Sell Opportunity"
    elif 20 <= latest_k <= 80 or 20 <= latest_d <= 80:
        return "Neutral Zone: Hold Phase"
    elif latest_k < 20 or latest_d < 20:
        return "Green Zone: Oversold - Potential Buy Opportunity"

# Function to decide action based on zone
def decide_action(zone):
    if "Overbought" in zone:
        return "Consider Selling"
    elif "Oversold" in zone:
        return "Consider Buying"
    else:
        return "Hold"

# Function to log stock analysis
def log_stock_analysis(stock_name, price, latest_k, latest_d, zone, decision, file_name="trade_log.xlsx"):
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Stock Analysis"]

    # Check if the stock already exists in the sheet
    found = False
    for row in ws.iter_rows(min_row=2):
        if row[0].value == stock_name:
            # Update existing row
            row[1].value = price
            row[2].value = latest_k
            row[3].value = latest_d
            row[4].value = zone
            row[5].value = decision
            found = True
            break

    if not found:
        # Append new row
        ws.append([stock_name, price, latest_k, latest_d, zone, decision])

    wb.save(file_name)

# Function to log American Bull information
def log_american_bull_info(stock_name, signal, date, file_name="trade_log.xlsx"):
    wb = openpyxl.load_workbook(file_name)
    ws = wb["American Bull Info"]

    # Check if the stock already exists in the sheet
    found = False
    for row in ws.iter_rows(min_row=2):
        if row[0].value == stock_name:
            # Update existing row
            row[1].value = signal
            row[2].value = date
            found = True
            break

    if not found:
        # Append new row
        ws.append([stock_name, signal, date])

    wb.save(file_name)

# Function to check if a trade already exists
def trade_exists(date, stock_name, action, file_name="trade_log.xlsx"):
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Trade Log"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_date, row_action, row_stock_name = row[0], row[1], row[2]
        if row_date == date and row_action == action and row_stock_name == stock_name:
            return True
    return False

# Function to log the buy or sell action
def log_trade(action, stock_name, price, shares, executed, zone, percent_value, file_name="trade_log.xlsx"):
    # Open the workbook and the worksheet
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Trade Log"]

    # Get today's date
    today_date = get_todays_date()

    # Search for existing entry
    found = False
    for row in ws.iter_rows(min_row=2):
        row_date = row[0].value
        row_action = row[1].value
        row_stock_name = row[2].value
        if row_date == today_date and row_action == action and row_stock_name == stock_name:
            # Update existing row
            row[3].value = price
            row[4].value = shares
            row[5].value = zone
            row[6].value = percent_value
            row[7].value = "Yes" if executed else "No"
            found = True
            break

    if not found:
        # Append new row
        executed_status = "Yes" if executed else "No"
        ws.append([today_date, action, stock_name, price, shares, zone, percent_value, executed_status])

    # Save the workbook
    wb.save(file_name)

# Function to get today's date in mm-dd format
def get_todays_date():
    return datetime.today().strftime('%m/%d')

# Buys $250 worth of stock
def order(stock_code, zone, percent_value):
    today_date = get_todays_date()
    if trade_exists(today_date, stock_code, "Buy"):
        print(f"Already bought {stock_code} today.")
        # Even if trade was not executed, log the attempt with Executed = "No"
        stock_price = float(stocks.get_latest_price(stock_code)[0])
        num_shares = 250 / stock_price
        log_trade("Buy", stock_code, stock_price, round(num_shares, 6), executed=False, zone=zone, percent_value=percent_value)
        return
    stock_price = float(stocks.get_latest_price(stock_code)[0])
    num_shares = 250 / stock_price

    # Place the buy order
    response = orders.order_buy_market(stock_code, round(num_shares, 6))
    print(f"Bought {round(num_shares, 6)} shares of {stock_code} at {stock_price}")

    # Log the trade in the Excel file
    log_trade("Buy", stock_code, stock_price, round(num_shares, 6), executed=True, zone=zone, percent_value=percent_value)

    return response

# Sells all stock
def sell(stock_code, zone, percent_value):
    today_date = get_todays_date()
    if trade_exists(today_date, stock_code, "Sell"):
        print(f"Already sold {stock_code} today.")
        # Even if trade was not executed, log the attempt with Executed = "No"
        holdings = account.build_holdings()
        if stock_code in holdings:
            shares_owned = float(holdings[stock_code]['quantity'])
        else:
            shares_owned = 0
        stock_price = float(stocks.get_latest_price(stock_code)[0])
        log_trade("Sell", stock_code, stock_price, shares_owned, executed=False, zone=zone, percent_value=percent_value)
        return

    # Fetch the holdings from the account module
    holdings = account.build_holdings()

    # Check if the stock code is in holdings
    if stock_code in holdings:
        shares_owned = float(holdings[stock_code]['quantity'])
        stock_price = float(stocks.get_latest_price(stock_code)[0])

        # Place the sell order
        response = orders.order_sell_market(stock_code, shares_owned)
        print(f"Sold all {shares_owned} shares of {stock_code} at {stock_price}")

        # Log the trade in the Excel file
        log_trade("Sell", stock_code, stock_price, shares_owned, executed=True, zone=zone, percent_value=percent_value)

        return response
    else:
        print(f"No shares of {stock_code} owned.")
        # Log the attempt with Executed = "No"
        stock_price = float(stocks.get_latest_price(stock_code)[0])
        log_trade("Sell", stock_code, stock_price, 0, executed=False, zone=zone, percent_value=percent_value)
        return None

# Function to retrieve owned stocks
def get_owned_stocks():
    holdings = account.build_holdings()
    stock_codes = list(holdings.keys())
    return stock_codes

# Initialize Excel to ensure it exists
initialize_excel()

# Credentials
username = 'your_username_here'
password = 'your_password_here'

# Initial login attempt
login = r.login(username, password)

portfolio_info = account.build_user_profile()
balance = portfolio_info['equity']

positions = account.build_holdings()

if positions:
    print("Here are the stocks you own:")
    for symbol, data in positions.items():
        print(f"Stock: {symbol}")
        print(f"    Quantity: {data['quantity']}")
        print(f"    Average Buy Price: {data['average_buy_price']}")
        print(f"    Current Price: {data['price']}")
        print(f"    Equity: {data['equity']}")
        print(f"    Percentage Change: {data['percent_change']}\n")
else:
    print("You do not own any stocks.")

# Main section for processing recommendations and executing trades
date = get_todays_date()
bull_table = recommmend_data()

predefined_stock_list = ['AAPL', 'MSFT', 'GOOG', 'AMZN', 'TSLA']  # Add your desired stocks here

owned_stocks = get_owned_stocks()
stock_list = list(set(owned_stocks + predefined_stock_list))  # Avoid duplicates

# Analyze each stock and make trading decisions
for stock in stock_list:
    stock_sanitized = stock.replace('.', '')

    # Get American Bull data
    bull_my_table = specifc_stock(stock_sanitized)
    if not bull_my_table or len(bull_my_table) == 0:
        print(f"No American Bull data for {stock_sanitized}")
        continue

    bull_date = bull_my_table[0]["Date"]
    bull_signal = bull_my_table[0]["Signal"]

    # Log the American Bull info
    log_american_bull_info(stock, bull_signal, bull_date)

    # Calculate stochastic oscillator
    latest_k, latest_d = calculate_stochastic(stock)
    if latest_k is None or latest_d is None:
        print(f"Could not calculate stochastic oscillator for {stock}")
        continue

    # Determine the zone
    zone = determine_zone(latest_k, latest_d)

    # Log the stock analysis
    latest_price = float(stocks.get_latest_price(stock)[0])
    decision = decide_action(zone)
    log_stock_analysis(stock, latest_price, latest_k, latest_d, zone, decision)

    today_date = get_todays_date()

    if str(bull_date) == today_date:
        
        if "BUY" in bull_signal and "Green Zone" in zone:
            order(stock, zone, latest_k)
        elif "SELL" in bull_signal and "Red Zone" in zone:
            sell(stock, zone, latest_k)
        else:
            print(bull_signal, f"Signal and zone do not match for {stock}")
    else:
        print(f"American Bull date {bull_date} does not match today's date for {stock}")

# Print account balance
print(f"Your account balance is: ${balance}")
